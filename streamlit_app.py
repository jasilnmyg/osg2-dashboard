import streamlit as st
import pandas as pd
from io import BytesIO
import re
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, PageBreak, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from collections import defaultdict
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import letter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import streamlit.components.v1 as components
from openpyxl.utils import get_column_letter
import time
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# If you also need these for more advanced PDF features:
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader


st.set_page_config(
    page_title="OSG DASHBOARD",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Force light mode and disable dark mode
st.markdown("""
<style>
/* Define light and dark themes explicitly */
html {
  --primary-light: #3498db;
  --secondary-light: #2980b9;
  --text-light: #2c3e50;
  --bg-light: #ffffff;
  --card-bg-light: #f8f9fa;
  --border-light: #dfe6e9;

  --primary-dark: #2980b9;
  --secondary-dark: #1c5d99;
  --text-dark: #ecf0f1;
  --bg-dark: #1e293b;
  --card-bg-dark: #334155;
  --border-dark: #475569;
}

/* Set default (light) mode */
body {
  background-color: var(--bg-light);
  color: var(--text-light);
}

/* Dark mode override */
@media (prefers-color-scheme: dark) {
  body {
    background-color: var(--bg-dark);
    color: var(--text-dark);
  }
}

/* Card Styling */
.report-card {
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
    border-left: 4px solid var(--primary-light);
    background-color: var(--card-bg-light);
}

@media (prefers-color-scheme: dark) {
    .report-card {
        background-color: var(--card-bg-dark);
        border-left: 4px solid var(--primary-dark);
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.4);
    }
}

/* Title Styling */
.report-title {
    font-size: 1.75rem;
    font-weight: 700;
    margin-bottom: 0.5rem;
    border-bottom: 2px solid var(--primary-light);
    padding-bottom: 0.5rem;
    color: var(--text-light);
}

@media (prefers-color-scheme: dark) {
    .report-title {
        color: var(--text-dark);
        border-bottom: 2px solid var(--primary-dark);
    }
}

/* Subtitle */
.report-subtitle {
    font-size: 1.25rem;
    font-weight: 600;
    margin: 1rem 0 0.5rem 0;
    color: var(--text-light);
}

@media (prefers-color-scheme: dark) {
    .report-subtitle {
        color: var(--text-dark);
    }
}

/* Time Indicator */
.time-indicator {
    display: inline-block;
    background-color: var(--primary-light);
    color: white;
    padding: 0.25rem 0.75rem;
    border-radius: 20px;
    font-size: 0.9rem;
    font-weight: 500;
}

@media (prefers-color-scheme: dark) {
    .time-indicator {
        background-color: var(--primary-dark);
    }
}

/* File Uploader */
.stFileUploader > div > div {
    border: 2px dashed var(--border-light);
    border-radius: 12px;
    padding: 2rem;
    background-color: var(--card-bg-light);
    transition: all 0.3s ease;
}

.stFileUploader > div > div:hover {
    border-color: var(--primary-light);
    background-color: rgba(52, 152, 219, 0.05);
}

@media (prefers-color-scheme: dark) {
    .stFileUploader > div > div {
        border: 2px dashed var(--border-dark);
        background-color: var(--card-bg-dark);
    }
    .stFileUploader > div > div:hover {
        border-color: var(--primary-dark);
        background-color: rgba(41, 128, 185, 0.1);
    }
}

/* Default File Message */
.default-file {
    font-size: 0.9rem;
    margin-top: 1rem;
    padding: 0.75rem;
    border-radius: 8px;
    border-left: 3px solid var(--primary-light);
    background-color: rgba(52, 152, 219, 0.1);
    color: var(--text-light);
}

@media (prefers-color-scheme: dark) {
    .default-file {
        background-color: rgba(41, 128, 185, 0.2);
        border-left: 3px solid var(--primary-dark);
        color: var(--text-dark);
    }
}
</style>
""", unsafe_allow_html=True)


# Neon glowing icons as SVG for tabs (can also use emojis or images)
tab_icons = {
    "📊 OSG REPORT 1": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M13 2h-2v10h2V2zM6 9h2v13H6V9zm10 0h2v13h-2V9z"/></svg>""",
    "📊 OSG REPORT 2": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M3 17h2v4H3v-4zm4-6h2v10H7V11zm4-4h2v14h-2V7zm4 6h2v8h-2v-8z"/></svg>""",
    "🔗 Data Mapping": """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" fill="currentColor"><path d="M15.5 14h-.79l-.28-.27A6.471 6.471 0 0016 9.5 6.5 6.5 0 109.5 16c1.61 0 3.09-.59 4.23-1.57l.27.28v.79l5 4.99L20.49 19l-4.99-5zM9.5 14C7.57 14 6 12.43 6 10.5S7.57 7 9.5 7 13 8.57 13 10.5 11.43 14 9.5 14z"/></svg>"""
}

# Streamlit Tabs with icons + neon styles
tab1, tab2, tab3 = st.tabs(list(tab_icons.keys()))


# --------------------------- REPORT 1 TAB ---------------------------
import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


# --------------------------- REPORT 1 TAB ---------------------------
with tab1:
    st.markdown('<h1 class="header">OSG All Store Report</h1>', unsafe_allow_html=True)

    st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following files to generate the sales summary report:
            <ul>
                <li><strong>Current Month sales Data</strong></li>
                <li><strong>Previous Month sales Data</strong></li>
                <li><strong>myG All Store List is loaded by default</strong></li>
                <li><strong>Store, RBM, BDM List is loaded by default</strong></li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        report_date = st.date_input("Select current report date", value=datetime.today())
    with col2:
        prev_date = st.date_input("Select previous report date (for comparison)", value=datetime.today().replace(day=1))

    book1_file = st.file_uploader("Upload current month sales data", type=["xlsx"], key="curr_sales")
    prev_month_file = st.file_uploader("Upload previous month sales data", type=["xlsx"], key="prev_sales")

    store_list_file = "Myg All Store.xlsx"
    rbm_bdm_file = "/workspaces/osg2-dashboard/RBM,BDM,BRANCH.xlsx"

    try:
        future_store_df = pd.read_excel(store_list_file)
        rbm_bdm_df = pd.read_excel(rbm_bdm_file)
        st.success("✅ Loaded default Future Store List & Store,RBM,BDM List.")
    except Exception as e:
        st.error(f"Error loading defaults: {e}")
        st.stop()

    if book1_file:
        book1_df = pd.read_excel(book1_file)
        book1_df.rename(columns={'Branch': 'Store'}, inplace=True)
        book1_df['DATE'] = pd.to_datetime(book1_df['DATE'], dayfirst=True, errors='coerce')
        book1_df = book1_df.dropna(subset=['DATE'])
        rbm_bdm_df.rename(columns={'Branch': 'Store'}, inplace=True)

        today = pd.to_datetime(report_date)
        mtd_df = book1_df[book1_df['DATE'].dt.month == today.month]
        today_df = mtd_df[mtd_df['DATE'].dt.date == today.date()]

        today_agg = today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'FTD Count', 'AMOUNT': 'FTD Amount'})
        mtd_agg = mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(columns={'QUANTITY': 'MTD Count', 'AMOUNT': 'MTD Amount'})

        if prev_month_file:
            prev_df = pd.read_excel(prev_month_file)
            prev_df.rename(columns={'Branch': 'Store'}, inplace=True)
            prev_df['DATE'] = pd.to_datetime(prev_df['DATE'], dayfirst=True, errors='coerce')
            prev_df = prev_df.dropna(subset=['DATE'])
            prev_month = pd.to_datetime(prev_date)
            prev_mtd_df = prev_df[prev_df['DATE'].dt.month == prev_month.month]
            prev_mtd_agg = prev_mtd_df.groupby('Store', as_index=False).agg({'AMOUNT': 'sum'}).rename(columns={'AMOUNT': 'PREV MONTH SALE'})
        else:
            prev_mtd_agg = pd.DataFrame(columns=['Store', 'PREV MONTH SALE'])

        all_stores = pd.DataFrame(pd.Series(pd.concat([future_store_df['Store'], book1_df['Store']]).unique(), name='Store'))
        report_df = all_stores.merge(today_agg, on='Store', how='left') \
                              .merge(mtd_agg, on='Store', how='left') \
                              .merge(prev_mtd_agg, on='Store', how='left') \
                              .merge(rbm_bdm_df[['Store', 'RBM', 'BDM']], on='Store', how='left')

        report_df[['FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']] = report_df[['FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']].fillna(0).astype(int)
        report_df['PREV MONTH SALE'] = report_df['PREV MONTH SALE'].fillna(0).astype(int)
        report_df['DIFF %'] = report_df.apply(
            lambda x: round(((x['MTD Amount'] - x['PREV MONTH SALE']) / x['PREV MONTH SALE']) * 100, 2) if x['PREV MONTH SALE'] != 0 else 0,
            axis=1
        )

        report_df['ASP'] = report_df.apply(
            lambda x: round(x['MTD Amount'] / x['MTD Count'], 2) if x['MTD Count'] != 0 else 0,
            axis=1
        )

        excel_output = BytesIO()
        with pd.ExcelWriter(excel_output, engine='xlsxwriter') as writer:
            workbook = writer.book

            colors_palette = {
                'primary_blue': '#1E3A8A',
                'light_blue': '#DBEAFE',
                'success_green': '#065F46',
                'light_green': '#D1FAE5',
                'warning_orange': '#EA580C',
                'light_orange': '#FED7AA',
                'danger_red': '#DC2626',
                'light_red': '#FEE2E2',
                'accent_purple': '#7C3AED',
                'light_purple': '#EDE9FE',
                'neutral_gray': '#6B7280',
                'light_gray': '#F9FAFB',
                'white': '#FFFFFF',
                'dark_blue': '#0F172A',
                'mint_green': '#10B981',
                'light_mint': '#ECFDF5',
                'royal_blue': '#3B82F6',
                'light_royal': '#EBF8FF'
            }

            formats = {
                'title': workbook.add_format({
                    'bold': True, 'font_size': 16, 'font_color': colors_palette['primary_blue'],
                    'align': 'center', 'valign': 'vcenter', 'bg_color': colors_palette['white'],
                    'border': 1, 'border_color': colors_palette['primary_blue']
                }),
                'subtitle': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['neutral_gray'],
                    'align': 'center', 'valign': 'vcenter', 'italic': True
                }),
                'header_main': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['primary_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['primary_blue'], 'text_wrap': True
                }),
                'header_secondary': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['primary_blue'],
                    'bg_color': colors_palette['light_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['primary_blue']
                }),
                'data_normal': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']
                }),
                'data_alternate': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray']
                }),
                'data_store_name': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1
                }),
                'data_store_name_alt': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_gray'], 'indent': 1
                }),
                'positive_value': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['success_green'], 'bg_color': colors_palette['light_green'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['success_green'], 'bold': True
                }),
                'negative_value': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'bold': True
                }),
                'zero_value': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['warning_orange'], 'bg_color': colors_palette['light_orange'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['warning_orange'], 'bold': True
                }),
                'total_row': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['accent_purple']
                }),
                'total_label': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['accent_purple']
                }),
                'rbm_title': workbook.add_format({
                    'bold': True, 'font_size': 18, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['dark_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['dark_blue']
                }),
                'rbm_subtitle': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['dark_blue'],
                    'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['royal_blue'], 'italic': True
                }),
                'rbm_header': workbook.add_format({
                    'bold': True, 'font_size': 11, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['royal_blue'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['royal_blue'], 'text_wrap': True
                }),
                'rbm_data_normal': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white']
                }),
                'rbm_data_alternate': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal']
                }),
                'rbm_store_name': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['white'], 'indent': 1
                }),
                'rbm_store_name_alt': workbook.add_format({
                    'font_size': 10, 'bold': True, 'align': 'left', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'indent': 1
                }),
                'rbm_positive': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['mint_green'], 'bg_color': colors_palette['light_mint'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['mint_green'], 'bold': True
                }),
                'rbm_negative': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['danger_red'], 'bg_color': colors_palette['light_red'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['danger_red'], 'bold': True
                }),
                'rbm_zero': workbook.add_format({
                    'font_size': 10, 'font_color': colors_palette['warning_orange'], 'bg_color': colors_palette['light_orange'],
                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': colors_palette['warning_orange'], 'bold': True
                }),
                'rbm_total': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green']
                }),
                'rbm_total_label': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green']
                }),
                'rbm_summary': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['royal_blue'],
                    'bg_color': colors_palette['light_royal'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['royal_blue']
                }),
                'rbm_performance': workbook.add_format({
                    'bold': True, 'font_size': 10, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['accent_purple'], 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['accent_purple']
                }),
                'asp_format': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'num_format': '₹#,##0.00'
                }),
                'asp_format_alt': workbook.add_format({
                    'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                    'border': 1, 'border_color': colors_palette['neutral_gray'], 'bg_color': colors_palette['light_royal'], 'num_format': '₹#,##0.00'
                }),
                'asp_total': workbook.add_format({
                    'bold': True, 'font_size': 12, 'font_color': colors_palette['white'],
                    'bg_color': colors_palette['mint_green'], 'align': 'center', 'valign': 'vcenter',
                    'border': 2, 'border_color': colors_palette['mint_green'], 'num_format': '₹#,##0.00'
                })
            }

            # ALL STORES SHEET
            all_data = report_df.sort_values('MTD Amount', ascending=False)
            worksheet = workbook.add_worksheet("All Stores")

            column_widths = [25, 12, 15, 12, 15]
            for i, width in enumerate(column_widths):
                worksheet.set_column(i, i, width)

            worksheet.merge_range(0, 0, 0, 4, "OSG All Stores Sales Report", formats['title'])
            worksheet.merge_range(1, 0, 1, 4, f"Report Generated: {datetime.now().strftime('%d %B %Y')}", formats['subtitle'])

            total_stores = len(all_data)
            active_stores = len(all_data[all_data['FTD Count'] > 0])
            inactive_stores = total_stores - active_stores

            worksheet.merge_range(3, 0, 3, 1, "📊 SUMMARY", formats['header_secondary'])
            worksheet.merge_range(3, 2, 3, 4, f"Total: {total_stores} | Active: {active_stores} | Inactive: {inactive_stores}", formats['data_normal'])

            headers = ['Store Name', 'FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']
            for col, header in enumerate(headers):
                worksheet.write(5, col, header, formats['header_main'])

            for row_idx, (_, row) in enumerate(all_data.iterrows(), start=6):
                is_alternate = (row_idx - 6) % 2 == 1
                store_format = formats['data_store_name_alt'] if is_alternate else formats['data_store_name']
                worksheet.write(row_idx, 0, row['Store'], store_format)

                ftd_count = int(row['FTD Count'])
                if ftd_count == 0:
                    worksheet.write(row_idx, 1, ftd_count, formats['zero_value'])
                else:
                    worksheet.write(row_idx, 1, ftd_count, formats['positive_value'])

                data_format = formats['data_alternate'] if is_alternate else formats['data_normal']
                worksheet.write(row_idx, 2, int(row['FTD Amount']), data_format)

                mtd_count = int(row['MTD Count'])
                if mtd_count == 0:
                    worksheet.write(row_idx, 3, mtd_count, formats['zero_value'])
                else:
                    worksheet.write(row_idx, 3, mtd_count, formats['positive_value'])

                worksheet.write(row_idx, 4, int(row['MTD Amount']), data_format)

            total_row = len(all_data) + 7
            worksheet.write(total_row, 0, '🎯 TOTAL', formats['total_label'])
            worksheet.write(total_row, 1, all_data['FTD Count'].sum(), formats['total_row'])
            worksheet.write(total_row, 2, all_data['FTD Amount'].sum(), formats['total_row'])
            worksheet.write(total_row, 3, all_data['MTD Count'].sum(), formats['total_row'])
            worksheet.write(total_row, 4, all_data['MTD Amount'].sum(), formats['total_row'])

            if len(all_data) > 0:
                top_performer = all_data.iloc[0]
                insights_row = total_row + 2
                worksheet.merge_range(insights_row, 0, insights_row, 4,
                                    f"🏆 Top Performer: {top_performer['Store']} (₹{int(top_performer['MTD Amount']):,})",
                                    formats['positive_value'])

            # RBM SHEETS
            for rbm in report_df['RBM'].dropna().unique():
                rbm_data = report_df[report_df['RBM'] == rbm].sort_values('MTD Amount', ascending=False)
                worksheet_name = rbm[:31] if len(rbm) > 31 else rbm
                rbm_ws = workbook.add_worksheet(worksheet_name)

                rbm_column_widths = [25, 12, 15, 12, 15, 15, 12, 12]
                for i, width in enumerate(rbm_column_widths):
                    rbm_ws.set_column(i, i, width)

                rbm_ws.merge_range(0, 0, 0, 7, f" {rbm} - Sales Performance Report", formats['rbm_title'])
                rbm_ws.merge_range(1, 0, 1, 7, f"Report Period: {datetime.now().strftime('%B %Y')} | Generated: {datetime.now().strftime('%d %B %Y')}", formats['rbm_subtitle'])

                rbm_total_stores = len(rbm_data)
                rbm_active_stores = len(rbm_data[rbm_data['FTD Count'] > 0])
                rbm_inactive_stores = rbm_total_stores - rbm_active_stores
                rbm_total_amount = rbm_data['MTD Amount'].sum()

                rbm_ws.merge_range(3, 0, 3, 1, "📈 PERFORMANCE OVERVIEW", formats['rbm_summary'])
                rbm_ws.merge_range(3, 2, 3, 7, f"Total Stores: {rbm_total_stores} | Active: {rbm_active_stores} | Inactive: {rbm_inactive_stores} | Total Revenue: ₹{rbm_total_amount:,}", formats['rbm_summary'])

                if len(rbm_data) > 0:
                    best_performer = rbm_data.iloc[0]
                    rbm_ws.merge_range(4, 0, 4, 7, f"🥇 Best Performer: {best_performer['Store']} - ₹{int(best_performer['MTD Amount']):,}", formats['rbm_performance'])

                headers = ['Store Name', 'FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount', 'PREV MONTH SALE', 'DIFF %', 'ASP']
                for col, header in enumerate(headers):
                    rbm_ws.write(6, col, header, formats['rbm_header'])

                for row_idx, (_, row) in enumerate(rbm_data.iterrows(), start=7):
                    is_alternate = (row_idx - 7) % 2 == 1
                    store_format = formats['rbm_store_name_alt'] if is_alternate else formats['rbm_store_name']
                    rbm_ws.write(row_idx, 0, row['Store'], store_format)

                    ftd_count = int(row['FTD Count'])
                    if ftd_count == 0:
                        rbm_ws.write(row_idx, 1, ftd_count, formats['rbm_zero'])
                    else:
                        rbm_ws.write(row_idx, 1, ftd_count, formats['rbm_positive'])

                    data_format = formats['rbm_data_alternate'] if is_alternate else formats['rbm_data_normal']
                    rbm_ws.write(row_idx, 2, int(row['FTD Amount']), data_format)

                    mtd_count = int(row['MTD Count'])
                    if mtd_count == 0:
                        rbm_ws.write(row_idx, 3, mtd_count, formats['rbm_zero'])
                    else:
                        rbm_ws.write(row_idx, 3, mtd_count, formats['rbm_positive'])

                    rbm_ws.write(row_idx, 4, int(row['MTD Amount']), data_format)
                    rbm_ws.write(row_idx, 5, int(row['PREV MONTH SALE']), data_format)

                    diff_pct = row['DIFF %']
                    if diff_pct > 0:
                        rbm_ws.write(row_idx, 6, f"{diff_pct}%", formats['rbm_positive'])
                    elif diff_pct < 0:
                        rbm_ws.write(row_idx, 6, f"{diff_pct}%", formats['rbm_negative'])
                    else:
                        rbm_ws.write(row_idx, 6, f"{diff_pct}%", formats['rbm_zero'])

                    asp_format = formats['asp_format_alt'] if is_alternate else formats['asp_format']
                    rbm_ws.write(row_idx, 7, row['ASP'], asp_format)

                total_row = len(rbm_data) + 8
                rbm_ws.write(total_row, 0, '🎯 TOTAL', formats['rbm_total_label'])
                rbm_ws.write(total_row, 1, rbm_data['FTD Count'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 2, rbm_data['FTD Amount'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 3, rbm_data['MTD Count'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 4, rbm_data['MTD Amount'].sum(), formats['rbm_total'])
                rbm_ws.write(total_row, 5, rbm_data['PREV MONTH SALE'].sum(), formats['rbm_total'])

                total_prev = rbm_data['PREV MONTH SALE'].sum()
                total_curr = rbm_data['MTD Amount'].sum()
                overall_growth = round(((total_curr - total_prev) / total_prev) * 100, 2) if total_prev != 0 else 0

                if overall_growth > 0:
                    rbm_ws.write(total_row, 6, f"{overall_growth}%", formats['rbm_total'])
                elif overall_growth < 0:
                    rbm_ws.write(total_row, 6, f"{overall_growth}%", formats['rbm_total'])
                else:
                    rbm_ws.write(total_row, 6, f"{overall_growth}%", formats['rbm_total'])

                total_mtd_count = rbm_data['MTD Count'].sum()
                total_mtd_amount = rbm_data['MTD Amount'].sum()
                overall_asp = round(total_mtd_amount / total_mtd_count, 2) if total_mtd_count != 0 else 0
                rbm_ws.write(total_row, 7, overall_asp, formats['asp_total'])

                insights_row = total_row + 2
                if overall_growth > 15:
                    rbm_ws.merge_range(insights_row, 0, insights_row, 7,
                                     f"📈 Excellent Growth: {overall_growth}% increase from previous month",
                                     formats['rbm_positive'])
                elif overall_growth < 0:
                    rbm_ws.merge_range(insights_row, 0, insights_row, 7,
                                     f"📉 Needs Attention: {abs(overall_growth)}% decrease from previous month",
                                     formats['rbm_negative'])
                else:
                    rbm_ws.merge_range(insights_row, 0, insights_row, 7,
                                     f"📊 Stable Performance: Less change from previous month",
                                     formats['rbm_zero'])

                insights_row += 1
                top_3_stores = rbm_data.head(3)
                if len(top_3_stores) > 0:
                    top_stores_text = " | ".join([f"{store['Store']}: ₹{int(store['MTD Amount']):,}"
                                                for _, store in top_3_stores.iterrows()])
                    rbm_ws.merge_range(insights_row, 0, insights_row, 7,
                                     f"🏆 Top 3 Performers: {top_stores_text}",
                                     formats['rbm_summary'])

            # BDM REPORT SHEET
            bdm_data = report_df.groupby('BDM').agg({
                'FTD Count': 'sum',
                'FTD Amount': 'sum',
                'MTD Count': 'sum',
                'MTD Amount': 'sum'
            }).reset_index().sort_values('MTD Amount', ascending=False)

            bdm_ws = workbook.add_worksheet("BDM Report")

            bdm_column_widths = [25, 12, 15, 12, 15]
            for i, width in enumerate(bdm_column_widths):
                bdm_ws.set_column(i, i, width)

            bdm_ws.merge_range(0, 0, 0, 4, "BDM Sales Performance Report", formats['rbm_title'])
            bdm_ws.merge_range(1, 0, 1, 4, f"Report Period: {datetime.now().strftime('%B %Y')} | Generated: {datetime.now().strftime('%d %B %Y')}", formats['rbm_subtitle'])

            bdm_total_bdms = len(bdm_data)
            bdm_active_bdms = len(bdm_data[bdm_data['FTD Count'] > 0])
            bdm_inactive_bdms = bdm_total_bdms - bdm_active_bdms
            bdm_total_amount = bdm_data['MTD Amount'].sum()

            bdm_ws.merge_range(3, 0, 3, 1, "📈 PERFORMANCE OVERVIEW", formats['rbm_summary'])
            bdm_ws.merge_range(3, 2, 3, 4, f"Total BDMs: {bdm_total_bdms} | Active: {bdm_active_bdms} | Inactive: {bdm_inactive_bdms} | Total Revenue: ₹{bdm_total_amount:,}", formats['rbm_summary'])

            if len(bdm_data) > 0:
                best_performer = bdm_data.iloc[0]
                bdm_ws.merge_range(4, 0, 4, 4, f"🥇 Best Performer: {best_performer['BDM']} - ₹{int(best_performer['MTD Amount']):,}", formats['rbm_performance'])

            headers = ['BDM Name', 'FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']
            for col, header in enumerate(headers):
                bdm_ws.write(6, col, header, formats['rbm_header'])

            for row_idx, (_, row) in enumerate(bdm_data.iterrows(), start=7):
                is_alternate = (row_idx - 7) % 2 == 1
                bdm_format = formats['rbm_store_name_alt'] if is_alternate else formats['rbm_store_name']
                bdm_ws.write(row_idx, 0, row['BDM'], bdm_format)

                ftd_count = int(row['FTD Count'])
                if ftd_count == 0:
                    bdm_ws.write(row_idx, 1, ftd_count, formats['rbm_zero'])
                else:
                    bdm_ws.write(row_idx, 1, ftd_count, formats['rbm_positive'])

                data_format = formats['rbm_data_alternate'] if is_alternate else formats['rbm_data_normal']
                bdm_ws.write(row_idx, 2, int(row['FTD Amount']), data_format)

                mtd_count = int(row['MTD Count'])
                if mtd_count == 0:
                    bdm_ws.write(row_idx, 3, mtd_count, formats['rbm_zero'])
                else:
                    bdm_ws.write(row_idx, 3, mtd_count, formats['rbm_positive'])

                bdm_ws.write(row_idx, 4, int(row['MTD Amount']), data_format)

            total_row = len(bdm_data) + 8
            bdm_ws.write(total_row, 0, '🎯 TOTAL', formats['rbm_total_label'])
            bdm_ws.write(total_row, 1, bdm_data['FTD Count'].sum(), formats['rbm_total'])
            bdm_ws.write(total_row, 2, bdm_data['FTD Amount'].sum(), formats['rbm_total'])
            bdm_ws.write(total_row, 3, bdm_data['MTD Count'].sum(), formats['rbm_total'])
            bdm_ws.write(total_row, 4, bdm_data['MTD Amount'].sum(), formats['rbm_total'])

            insights_row = total_row + 2
            top_3_bdms = bdm_data.head(3)
            if len(top_3_bdms) > 0:
                top_bdms_text = " | ".join([f"{bdm['BDM']}: ₹{int(bdm['MTD Amount']):,}"
                                          for _, bdm in top_3_bdms.iterrows()])
                bdm_ws.merge_range(insights_row, 0, insights_row, 4,
                                 f"🏆 Top 3 Performers: {top_bdms_text}",
                                 formats['rbm_summary'])

        excel_output.seek(0)
        st.success("✅ Excel report generated successfully!")
        st.download_button(
            label="📥 Download Detailed Excel Report",
            data=excel_output.getvalue(),
            file_name=f"OSG_Sales_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Click to download the comprehensive sales report with all RBM and BDM sheets"
        )

st.markdown("""
    <style>
    .insight-box {
        background-color: #f8f9fa;
        padding: 15px;
        border-radius: 10px;
        border-left: 4px solid #007bff;
        margin: 10px 0;
    }
    .insight-box h4 {
        color: #007bff;
        margin-top: 0;
    }
    .insight-box ul {
        margin-bottom: 0;
    }
    .insight-box li {
        margin: 5px 0;
    }
    </style>
""", unsafe_allow_html=True)

# --------------------------- REPORT 2 TAB ---------------------------
with tab2:
    st.markdown('<h1 class="header">OSG Day View Report</h1>', unsafe_allow_html=True)

    with st.container():
        st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following file to generate the store summary report:
            <ul>
                <li><strong>Daily Sales Report</strong></li>
                <li><strong>myG Future Store List</strong> is loaded by default</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    # Date and Time selection
    selected_date = st.date_input("Select Date", value=datetime.today())
    time_slot = st.selectbox("Select Time Slot", options=["12:30PM", "1PM", "4PM", "6PM"])
    formatted_date = selected_date.strftime("%d-%m-%Y")
    report_title = f"{formatted_date} EW Sale Till {time_slot}"

    # File uploader for sales report
    book2_file = st.file_uploader("Upload Daily Sales Report", type=["xlsx"], key="r2_book1")

    # Load Future Store List
    future_df = pd.read_excel("/workspaces/osg2-dashboard/Future Store List.xlsx")
    st.success("✅ Loaded default Future Store List.")

    if book2_file:
        with st.spinner('Processing data...'):
            book2_df = pd.read_excel(book2_file)
            book2_df.rename(columns={'Branch': 'Store'}, inplace=True)

            agg = book2_df.groupby('Store', as_index=False).agg({
                'QUANTITY': 'sum',
                'AMOUNT': 'sum'
            })

            all_stores = pd.DataFrame(pd.concat([future_df['Store'], agg['Store']]).unique(), columns=['Store'])
            merged = all_stores.merge(agg, on='Store', how='left')
            merged['QUANTITY'] = merged['QUANTITY'].fillna(0).astype(int)
            merged['AMOUNT'] = merged['AMOUNT'].fillna(0).astype(int)

            merged = merged.sort_values(by='AMOUNT', ascending=False).reset_index(drop=True)

            total = pd.DataFrame([{
                'Store': 'TOTAL',
                'QUANTITY': merged['QUANTITY'].sum(),
                'AMOUNT': merged['AMOUNT'].sum()
            }])

            final_df = pd.concat([merged, total], ignore_index=True)
            final_df.rename(columns={'Store': 'Branch'}, inplace=True)

            # Excel report generator
            def generate_report2_excel(df, title_text):
                wb = Workbook()
                ws = wb.active
                ws.title = "Store Report"

                # Title
                ws.merge_cells('A1:C1')
                title_cell = ws['A1']
                title_cell.value = title_text
                title_cell.font = Font(bold=True, size=11, color="FFFFFF")
                title_cell.alignment = Alignment(horizontal='center')
                title_cell.fill = PatternFill("solid", fgColor="4F81BD")

                # Styles
                header_fill = PatternFill("solid", fgColor="4F81BD")
                data_fill = PatternFill("solid", fgColor="DCE6F1")
                red_fill = PatternFill("solid", fgColor="F4CCCC")
                total_fill = PatternFill("solid", fgColor="FFD966")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
                header_font = Font(bold=True, color="FFFFFF")
                bold_font = Font(bold=True)

                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)

                        if r_idx == 2:
                            cell.fill = header_fill
                            cell.font = header_font
                        elif df.loc[r_idx - 3, 'Branch'] == 'TOTAL':
                            cell.fill = total_fill
                            cell.font = bold_font
                        elif df.loc[r_idx - 3, 'AMOUNT'] <= 0:
                            cell.fill = red_fill
                        else:
                            cell.fill = data_fill

                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                # Adjust column widths
                for col_idx, column_cells in enumerate(ws.columns, start=1):
                    max_length = 0
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer

            excel_buf2 = generate_report2_excel(final_df, report_title)

        with st.container():
            st.download_button(
                label="📥 Download Store Summary Report",
                data=excel_buf2,
                file_name=f"Store_Summary_{formatted_date}_{time_slot}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download store summary report in Excel format"
            )
    else:
        st.info("ℹ️ Please upload the Daily Sales Report to generate the store summary.")
# --------------------------- REPORT 3 TAB ---------------------------
with tab3:
    st.markdown('<h1 class="header">OSG & Product Data Mapping</h1>', unsafe_allow_html=True)

    with st.container():
        st.markdown("""
        <div class="info-box">
            <strong>Instructions:</strong> Upload the following files to map OSG and product data:
            <ul>
                <li><strong>OSG File</strong> - Contains warranty and protection plan data</li>
                <li><strong>PRODUCT File</strong> - Contains product information including models, categories, and IMEIs</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    # File upload section
    with st.container():
        st.markdown('<div class="file-upload-section">', unsafe_allow_html=True)
        osg_file = st.file_uploader(
            "Upload OSG File",
            type=["xlsx"],
            key="osg_mapping"
        )
        product_file = st.file_uploader(
            "Upload PRODUCT File",
            type=["xlsx"],
            key="product_mapping"
        )
        st.markdown('</div>', unsafe_allow_html=True)

    if osg_file and product_file:
        with st.spinner('Mapping data...'):
            osg_df = pd.read_excel(osg_file)
            product_df = pd.read_excel(product_file)

            # SKU Mapping
            sku_category_mapping = {
                "Warranty : Water Cooler/Dispencer/Geyser/RoomCooler/Heater": [
                    "COOLER", "DISPENCER", "GEYSER", "ROOM COOLER", "HEATER", "WATER HEATER", "WATER DISPENSER"
                ],
                "Warranty : Fan/Mixr/IrnBox/Kettle/OTG/Grmr/Geysr/Steamr/Inductn": [
                    "FAN", "MIXER", "IRON BOX", "KETTLE", "OTG", "GROOMING KIT", "GEYSER", "STEAMER", "INDUCTION",
                    "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "INDUCTION COOKER", "ELECTRIC KETTLE", "WALL FAN", "MIXER GRINDER", "CELLING FAN"
                ],
                "AC : EWP : Warranty : AC": ["AC", "AIR CONDITIONER", "AC INDOOR"],
                "HAEW : Warranty : Air Purifier/WaterPurifier": ["AIR PURIFIER", "WATER PURIFIER"],
                "HAEW : Warranty : Dryer/MW/DishW": ["DRYER", "MICROWAVE OVEN", "DISH WASHER", "MICROWAVE OVEN-CONV"],
                "HAEW : Warranty : Ref/WM": [
                    "REFRIGERATOR", "WASHING MACHINE", "WASHING MACHINE-TL", "REFRIGERATOR-DC",
                    "WASHING MACHINE-FL", "WASHING MACHINE-SA", "REF", "REFRIGERATOR-CBU", "REFRIGERATOR-FF", "WM"
                ],
                "HAEW : Warranty : TV": ["TV", "TV 28 %", "TV 18 %"],
                "TV : TTC : Warranty and Protection : TV": ["TV", "TV 28 %", "TV 18 %"],
                "TV : Spill and Drop Protection": ["TV", "TV 28 %", "TV 18 %"],
                "HAEW : Warranty :Chop/Blend/Toast/Air Fryer/Food Processr/JMG/Induction": [
                    "CHOPPER", "BLENDER", "TOASTER", "AIR FRYER", "FOOD PROCESSOR", "JUICER", "INDUCTION COOKER"
                ],
                "HAEW : Warranty : HOB and Chimney": ["HOB", "CHIMNEY"],
                "HAEW : Warranty : HT/SoundBar/AudioSystems/PortableSpkr": [
                    "HOME THEATRE", "AUDIO SYSTEM", "SPEAKER", "SOUND BAR", "PARTY SPEAKER"
                ],
                "HAEW : Warranty : Vacuum Cleaner/Fans/Groom&HairCare/Massager/Iron": [
                    "VACUUM CLEANER", "FAN", "MASSAGER", "IRON BOX", "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "WALL FAN", "ROBO VACCUM CLEANER"
                ],
                "AC AMC": ["AC", "AC INDOOR"]
            }

            product_df['Category'] = product_df['Category'].str.upper().fillna('')
            product_df['Model'] = product_df['Model'].fillna('')
            product_df['Customer Mobile'] = product_df['Customer Mobile'].astype(str)
            product_df['Invoice Number'] = product_df['Invoice Number'].astype(str)
            product_df['Item Rate'] = pd.to_numeric(product_df['Item Rate'], errors='coerce')
            product_df['IMEI'] = product_df['IMEI'].astype(str).fillna('')
            product_df['Brand'] = product_df['Brand'].fillna('')
            osg_df['Customer Mobile'] = osg_df['Customer Mobile'].astype(str)

            def extract_price_slab(text):
                match = re.search(r"Slab\s*:\s*(\d+)K-(\d+)K", str(text))
                if match:
                    return int(match.group(1)) * 1000, int(match.group(2)) * 1000
                return None, None

            def get_model(row):
                mobile = row['Customer Mobile']
                retailer_sku = str(row['Retailer SKU'])
                invoice = str(row.get('Invoice Number', ''))
                user_products = product_df[product_df['Customer Mobile'] == mobile]

                if user_products.empty:
                    return ''
                unique_models = user_products['Model'].dropna().unique()
                if len(unique_models) == 1:
                    return unique_models[0]

                mapped_keywords = []
                for sku_key, keywords in sku_category_mapping.items():
                    if sku_key in retailer_sku:
                        mapped_keywords = [kw.lower() for kw in keywords]
                        break

                filtered = user_products[user_products['Category'].str.lower().isin(mapped_keywords)]
                if filtered['Model'].nunique() == 1:
                    return filtered['Model'].iloc[0]

                slab_min, slab_max = extract_price_slab(retailer_sku)
                if slab_min and slab_max:
                    slab_filtered = filtered[(filtered['Item Rate'] >= slab_min) & (filtered['Item Rate'] <= slab_max)]
                    if slab_filtered['Model'].nunique() == 1:
                        return slab_filtered['Model'].iloc[0]
                    invoice_filtered = slab_filtered[slab_filtered['Invoice Number'].astype(str) == invoice]
                    if invoice_filtered['Model'].nunique() == 1:
                        return invoice_filtered['Model'].iloc[0]

                return ''

            osg_df['Model'] = osg_df.apply(get_model, axis=1)
            category_brand_df = product_df[['Customer Mobile', 'Model', 'Category', 'Brand']].drop_duplicates()
            osg_df = osg_df.merge(category_brand_df, on=['Customer Mobile', 'Model'], how='left')

            invoice_pool = defaultdict(list)
            itemrate_pool = defaultdict(list)
            imei_pool = defaultdict(list)

            for _, row in product_df.iterrows():
                key = (row['Customer Mobile'], row['Model'])
                invoice_pool[key].append(row['Invoice Number'])
                itemrate_pool[key].append(row['Item Rate'])
                imei_pool[key].append(row['IMEI'])

            invoice_usage_counter = defaultdict(int)
            itemrate_usage_counter = defaultdict(int)
            imei_usage_counter = defaultdict(int)

            def assign_from_pool(row, pool, counter_dict):
                key = (row['Customer Mobile'], row['Model'])
                values = pool.get(key, [])
                index = counter_dict[key]
                if index < len(values):
                    counter_dict[key] += 1
                    return values[index]
                return ''

            osg_df['Product Invoice Number'] = osg_df.apply(lambda row: assign_from_pool(row, invoice_pool, invoice_usage_counter), axis=1)
            osg_df['Item Rate'] = osg_df.apply(lambda row: assign_from_pool(row, itemrate_pool, itemrate_usage_counter), axis=1)
            osg_df['IMEI'] = osg_df.apply(lambda row: assign_from_pool(row, imei_pool, imei_usage_counter), axis=1)
            osg_df['Store Code'] = osg_df['Product Invoice Number'].astype(str).apply(
                lambda x: re.search(r'\b([A-Z]{2,})\b', x).group(1) if re.search(r'\b([A-Z]{2,})\b', x) else ''
            )

            def extract_warranty_duration(sku):
                sku = str(sku)
                match = re.search(r'Dur\s*:\s*(\d+)\+(\d+)', sku)
                if match:
                    return int(match.group(1)), int(match.group(2))
                match = re.search(r'(\d+)\+(\d+)\s*SDP-(\d+)', sku)
                if match:
                    return int(match.group(1)), f"{match.group(3)}P+{match.group(2)}W"
                match = re.search(r'Dur\s*:\s*(\d+)', sku)
                if match:
                    return 1, int(match.group(1))
                match = re.search(r'(\d+)\+(\d+)', sku)
                if match:
                    return int(match.group(1)), int(match.group(2))
                return '', ''

            osg_df[['Manufacturer Warranty', 'Duration (Year)']] = osg_df['Retailer SKU'].apply(
                lambda sku: pd.Series(extract_warranty_duration(sku))
            )

            def highlight_row(row):
                missing_fields = pd.isna(row.get('Model')) or str(row.get('Model')).strip() == ''
                missing_fields |= pd.isna(row.get('IMEI')) or str(row.get('IMEI')).strip() == ''
                try:
                    if float(row.get('Plan Price', 0)) < 0:
                        missing_fields |= True
                except:
                    missing_fields |= True
                return ['background-color: lightblue'] * len(row) if missing_fields else [''] * len(row)

            final_columns = [
                'Customer Mobile', 'Date', 'Invoice Number','Product Invoice Number', 'Customer Name', 'Store Code', 'Branch', 'Region',
                'IMEI', 'Category', 'Brand', 'Quantity', 'Item Code', 'Model', 'Plan Type', 'EWS QTY', 'Item Rate',
                'Plan Price', 'Sold Price', 'Email', 'Product Count', 'Manufacturer Warranty', 'Retailer SKU', 'OnsiteGo SKU',
                'Duration (Year)', 'Total Coverage', 'Comment', 'Return Flag', 'Return against invoice No.',
                'Primary Invoice No.'
            ]

            for col in final_columns:
                if col not in osg_df.columns:
                    osg_df[col] = ''
            osg_df['Quantity'] = 1
            osg_df['EWS QTY'] = 1
            osg_df = osg_df[final_columns]

            st.markdown("""
            <div class="success-box">
                <strong>✅ Data Mapping Completed Successfully</strong>
                <p>The OSG and product data has been successfully mapped. You can now download the report.</p>
            </div>
            """, unsafe_allow_html=True)

            @st.cache_data
            def convert_df(df):
               output = io.BytesIO()
               styled_df = df.style.apply(highlight_row, axis=1)
               with pd.ExcelWriter(output, engine='openpyxl') as writer:
                styled_df.to_excel(writer, index=False)
               output.seek(0)
               return output

            excel_data = convert_df(osg_df)

        # Download section
        with st.container():
            st.markdown('<div class="download-section">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Download Mapped Data Report",
                data=excel_data,
                file_name="OSG_Product_Mapping_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download the mapped OSG and product data in Excel format"
            )
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("ℹ️ Please upload both required files to perform data mapping.")
